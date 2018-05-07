#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""Funciones auxiliares"""

from __future__ import unicode_literals
from __future__ import print_function
from __future__ import with_statement

import os
import shutil
from openpyxl import load_workbook
import zipfile
import datetime
import time
import urlparse
import yaml
import logging
import logging.config
import yaml
import download

from paths import CONFIG_DOWNLOADS_PATH
from paths import CATALOGS_INDEX_PATH, CONFIG_GENERAL_PATH

FREQ_ISO_TO_HUMAN = {
    "R/P1Y": "anual",
    "R/P6M": "semestral",
    "R/P3M": "trimestral",
    "R/P1M": "mensual",
    "R/P1D": "diaria"
}

SEPARATOR_WIDTH = 60


def safe_sheet_name(string):
    invalid_chars = "[]:*?/\\"
    for invalid_char in invalid_chars:
        string = string.replace(invalid_char, "_")
    return string


def indicators_to_text(simple_dict):
    text = "\n" + "\n".join(
        "{}: {}".format(key.ljust(40), value)
        for key, value in sorted(simple_dict.items(), key=lambda x: x[0])
    )
    return text


def timeit(method):

    def timed(*args, **kw):
        ts = time.time()
        result = method(*args, **kw)
        te = time.time()

        print('{} ({}, {}) {%2.2f} sec'.format(
            method.__name__, args, kw, te - ts)
        )
        return result

    return timed


def print_zipfile_info(path):
    zf = zipfile.ZipFile(path)
    for info in zf.infolist():
        print(info.filename)
        print('\tComment:\t', info.comment)
        print('\tModified:\t', datetime.datetime(*info.date_time))
        print('\tSystem:\t\t', info.create_system, '(0 = Windows, 3 = Unix)')
        print('\tZIP version:\t', info.create_version)
        print('\tCompressed:\t', info.compress_size, 'bytes')
        print('\tUncompressed:\t', info.file_size, 'bytes')


def compress_file(from_path, to_path):
    zf = zipfile.ZipFile(to_path, 'w', zipfile.ZIP_DEFLATED)
    try:
        zf.write(from_path, os.path.basename(from_path))
    finally:
        zf.close()
    # print_zipfile_info(to_path)


def freq_iso_to_xlseries(freq_iso8601):
    frequencies_map = {
        "R/P1Y": "Y",
        "R/P6M": "S",
        "R/P3M": "Q",
        "R/P1M": "M",
        "R/P1D": "D"
    }
    return frequencies_map[freq_iso8601]


def freq_iso_to_pandas(freq_iso8601, how="start"):
    frequencies_map_start = {
        "R/P1Y": "AS",
        "R/P6M": "6MS",
        "R/P3M": "QS",
        "R/P1M": "MS",
        "R/P1D": "DS"
    }
    frequencies_map_end = {
        "R/P1Y": "A",
        "R/P6M": "6M",
        "R/P3M": "Q",
        "R/P1M": "M",
        "R/P1D": "D"
    }
    if how == "start":
        return frequencies_map_start[freq_iso8601]
    elif how == "end":
        return frequencies_map_end[freq_iso8601]
    else:
        raise Exception(
            "{} no se reconoce para 'how': debe ser 'start' o 'end'".format(
                how))


def remove_other_files(directory):
    """Se asegura de que un directorio exista."""
    ensure_dir_exists(directory)
    shutil.rmtree(directory)
    ensure_dir_exists(directory)


def ensure_dir_exists(directory):
    """Se asegura de que un directorio exista."""
    if not os.path.exists(directory):
        os.makedirs(directory)


def get_ws_case_insensitive(wb, title):
    """Devuelve una hoja en un workbook sin importar mayúsculas/minúsculas."""
    return wb[find_ws_name(wb, title)]


def find_ws_name(wb, name):
    """Busca una hoja en un workbook sin importar mayúsculas/minúsculas."""
    if type(wb) == str or type(wb) == unicode:
        wb = load_workbook(wb, read_only=True, data_only=True)

    for sheetname in wb.sheetnames:
        if sheetname.lower() == name.lower():
            return sheetname

    return None


def row_from_cell_coord(coord):
    return int(filter(lambda x: x.isdigit(), coord))


def load_yaml(path):
    with open(path) as config_file:
        return yaml.load(config_file)


def get_catalogs_index():
    return load_yaml(CATALOGS_INDEX_PATH)


def list_catalogs():
    index = get_catalogs_index()
    for catalog in index:
        print('{} -> ({})'.format(catalog, index[catalog]['url']))


def get_general_config():
    return load_yaml(CONFIG_GENERAL_PATH)


def get_logger(name=__name__):
    levels = {
        'critical': logging.CRITICAL,
        'error': logging.ERROR,
        'warning': logging.WARNING,
        'info': logging.INFO,
        'debug': logging.DEBUG
    }

    logger = logging.getLogger(name)

    if 'TESTING' in os.environ:
        logger.disabled = True
        selected_level = logging.DEBUG
    else:
        config = get_general_config()
        selected_level = levels[config['logging']]
        logger.setLevel(selected_level)

    ch = logging.StreamHandler()
    ch.setLevel(selected_level)

    logging_formatter = logging.Formatter(
        '%(asctime)s - %(name)s - %(levelname)s - %(message)s', '%Y-%m-%d %H:%M:%S')
    ch.setFormatter(logging_formatter)
    logger.addHandler(ch)

    return logger

logger = get_logger(os.path.basename(__file__))


def print_log_separator(logger, message):
    logger.info("=" * SEPARATOR_WIDTH)
    logger.info("|" + " " * (SEPARATOR_WIDTH - 2) + "|")
    
    logger.info("|" + message.center(SEPARATOR_WIDTH - 2) + "|")
    
    logger.info("|" + " " * (SEPARATOR_WIDTH - 2) + "|")
    logger.info("=" * SEPARATOR_WIDTH)


def is_http_or_https(url):
    return urlparse.urlparse(url).scheme in ["http", "https"]


def get_catalog_download_config(catalog_id):
    try:
        configs = load_yaml(CONFIG_DOWNLOADS_PATH)
    except:
        logger.warning("No se pudo cargar el archivo de configuración 'config_downloads.yaml'.")
        logger.warning("Utilizando configuración default...")
        configs = {
            "defaults": {}
        }

    default_config = configs["defaults"]

    config = configs[catalog_id] if catalog_id in configs else {}
    if "catalog" not in config:
        config["catalog"] = {}
    if "sources" not in config:
        config["sources"] = {}

    for key, value in default_config.items():
        for subconfig in config.values():
            if key not in subconfig:
                subconfig[key] = value

    return config


def download_with_config(url, file_path, config):
    download.download_to_file(url, file_path, **config)
